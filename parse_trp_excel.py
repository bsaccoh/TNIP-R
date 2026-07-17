"""
TRP Drive Test Parser - Excel Output
=====================================
Parses TEMS Pocket .trp files and saves all records to a formatted Excel workbook.
KPI thresholds are fetched from the NTNIP system's DT Config (drive_test_config table).

HOW TO USE:
  1. Set TRP_FILE below to your .trp file path
  2. Set OUTPUT_EXCEL to where you want the output saved
  3. Make sure the NTNIP backend is running (default: http://localhost:4000)
  4. Run:  python parse_trp_excel.py
  5. Open the .xlsx file

Requires: pip install openpyxl requests
"""

import struct, zlib, zipfile, re, math, json
from datetime import datetime, timezone
from pathlib import Path

# ============================================================================
#   CONFIGURE THESE TWO PATHS
# ============================================================================

TRP_FILE = r"C:\Users\Saccoh1629182\Documents\Babah\BS\OCS\project\babah\gov\NTNIP\drive_test_logs\Orange_LTE_UL_Corrected_GPS_Orange_3GUL_Corrected_GPS_20260119T155413Z.trp"

OUTPUT_EXCEL = r"C:\Users\Saccoh1629182\Documents\Babah\BS\OCS\project\babah\gov\NTNIP\drive_test_output.xlsx"

API_BASE = "http://localhost:4000/api"

# ============================================================================


# ---------- Protobuf wire-format decoder ----------

def decode_varint(buf, pos):
    result, shift = 0, 0
    while pos < len(buf):
        b = buf[pos]; pos += 1
        result |= (b & 0x7F) << shift
        if not (b & 0x80):
            break
        shift += 7
        if shift > 49:
            break
    return result, pos

def parse_pb(buf):
    fields = {}
    p = 0
    while p < len(buf):
        tag, p = decode_varint(buf, p)
        fn = tag >> 3
        wt = tag & 7
        if fn == 0:
            break
        if wt == 0:
            val, p = decode_varint(buf, p)
            fields.setdefault(fn, []).append(('v', val))
        elif wt == 2:
            length, p = decode_varint(buf, p)
            if p + length > len(buf):
                break
            fields.setdefault(fn, []).append(('b', buf[p:p+length]))
            p += length
        elif wt == 1:
            if p + 8 > len(buf):
                break
            val = struct.unpack_from('<d', buf, p)[0]
            fields.setdefault(fn, []).append(('d', val))
            p += 8
        elif wt == 5:
            if p + 4 > len(buf):
                break
            val = struct.unpack_from('<f', buf, p)[0]
            fields.setdefault(fn, []).append(('f', val))
            p += 4
        else:
            break
    return fields

def iter_records(decompressed):
    records = []
    pos = 0
    while pos < len(decompressed):
        size, pos = decode_varint(decompressed, pos)
        if size <= 0 or size > 500000 or pos + size > len(decompressed):
            break
        records.append(parse_pb(decompressed[pos:pos+size]))
        pos += size
    return records

def decompress_cdf(raw):
    return zlib.decompress(raw[8:])


# ---------- Parameter IDs ----------

PID = {
    'RSRP': 9990, 'RSRQ': 9982, 'SINR': 12238, 'PCI': 10078,
    'BAND': 12390, 'DL_EARFCN': 13134, 'UL_EARFCN': 10190,
    'PDSCH_TP': 12214, 'PUSCH_TP': 10782,
    'PDSCH_TOTAL': 14454, 'PUSCH_TOTAL': 14448,
    'LATITUDE': 603, 'LONGITUDE': 602, 'SPEED': 600, 'ALTITUDE': 601,
}

def extract_value(sub):
    for t, v in sub.get(10, []):
        if t == 'f': return round(v, 4)
    for t, v in sub.get(11, []):
        if t == 'd': return round(v, 6)
    for t, v in sub.get(9, []):
        if t == 'v': return v
    for t, v in sub.get(8, []):
        if t == 'v': return v
    for t, v in sub.get(6, []):
        if t == 'v': return v
    return None


# ---------- XML metadata ----------

def xml_tag(xml, tag):
    m = re.search(rf'<{tag}[^>]*>([^<]*)</{tag}>', xml, re.I)
    return m.group(1).strip() if m else None

def xml_named_prop(xml, name):
    m = re.search(rf'<Property>\s*<Name>{name}</Name>\s*<Value[^>]*>([^<]*)</Value>', xml, re.I)
    return m.group(1).strip() or None if m else None

def parse_metadata(z):
    meta = {'device': None, 'os': None, 'app_version': None,
            'operator': None, 'mcc': None, 'mnc': None, 'imei': None,
            'start_time': None, 'tags': None}
    try:
        content_xml = z.read('trp/content.xml').decode('utf-8', errors='replace')
        meta['start_time'] = xml_tag(content_xml, 'Time') or xml_tag(content_xml, 'a:Time')
        meta['tags'] = xml_tag(content_xml, 'Tags')
        ver_major = xml_tag(content_xml, 'a:_Major')
        ver_minor = xml_tag(content_xml, 'a:_Minor')
        if ver_major:
            meta['app_version'] = f"TEMS Pocket {ver_major}.{ver_minor or '0'}"
    except Exception:
        pass
    try:
        sys_xml = z.read('trp/systeminformation.xml').decode('utf-8', errors='replace')
        mfr = xml_tag(sys_xml, 'Manufacturer')
        model = xml_tag(sys_xml, 'Model')
        if model:
            meta['device'] = f"{mfr} {model}" if mfr else model
        os_name = xml_tag(sys_xml, 'Caption')
        os_major = xml_tag(sys_xml, 'a:_Major')
        if os_name:
            meta['os'] = f"{os_name} {os_major}" if os_major and os_major != '-1' else os_name
    except Exception:
        pass
    try:
        sp_xml = z.read('trp/providers/sp1/serviceprovider.xml').decode('utf-8', errors='replace')
        meta['imei'] = xml_named_prop(sp_xml, 'IMEI')
        imsi = xml_named_prop(sp_xml, 'IMSI')
        if imsi and len(imsi) >= 5:
            meta['mcc'] = imsi[:3]
            meta['mnc'] = imsi[3:5]
    except Exception:
        pass
    if meta['mcc'] == '619':
        ops = {'01': 'Orange SL', '03': 'Africell SL', '05': 'Qcell SL', '04': 'Sierra Tel'}
        meta['operator'] = ops.get(meta['mnc'])
    return meta


# ---------- Haversine ----------

def haversine(lat1, lon1, lat2, lon2):
    R = 6371
    dLat = math.radians(lat2 - lat1)
    dLon = math.radians(lon2 - lon1)
    a = math.sin(dLat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dLon/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# ---------- Fetch KPI config from system ----------

DEFAULT_CONFIG = {
    'rsrp_threshold': -100, 'rsrq_threshold': -15, 'sinr_threshold': 0,
    'dl_threshold': 2000, 'ul_threshold': 500, 'coverage_target': 95,
    'rsrp_excellent': -80, 'rsrp_good': -90, 'rsrp_fair': -100, 'rsrp_poor': -110,
    'sinr_excellent': 20, 'sinr_good': 10, 'sinr_fair': 0,
    'score_weight_rsrp': 0.35, 'score_weight_sinr': 0.25,
    'score_weight_dl': 0.25, 'score_weight_rsrq': 0.15,
}

def fetch_config():
    try:
        import requests
        resp = requests.get(f"{API_BASE}/drive-tests/config", timeout=5)
        if resp.status_code == 200:
            data = resp.json().get('data', resp.json())
            merged = {**DEFAULT_CONFIG, **{k: v for k, v in data.items() if v is not None}}
            print(f"  Config loaded from {API_BASE}/drive-tests/config")
            return merged
    except Exception as e:
        print(f"  Could not fetch config from API ({e}), using defaults")
    return DEFAULT_CONFIG.copy()


# ---------- Coverage classification ----------

def rsrp_class(val, cfg):
    if val is None: return None
    if val >= cfg['rsrp_excellent']: return 'Excellent'
    if val >= cfg['rsrp_good']: return 'Good'
    if val >= cfg['rsrp_fair']: return 'Fair'
    if val >= cfg['rsrp_poor']: return 'Poor'
    return 'No Signal'

def sinr_class(val, cfg):
    if val is None: return None
    if val >= cfg['sinr_excellent']: return 'Excellent'
    if val >= cfg['sinr_good']: return 'Good'
    if val >= cfg['sinr_fair']: return 'Fair'
    return 'Poor'


# ---------- Main parser ----------

def parse_trp(filepath, cfg=None):
    if cfg is None:
        cfg = DEFAULT_CONFIG
    with zipfile.ZipFile(filepath) as z:
        meta = parse_metadata(z)

        provider_dirs = []
        for name in z.namelist():
            m = re.match(r'^trp/providers/(sp\d+)/$', name)
            if m:
                provider_dirs.append(m.group(1))
        if not provider_dirs:
            provider_dirs = ['sp1']

        all_param_names = {}
        samples = []

        for sp in provider_dirs:
            try:
                decl_raw = z.read(f'trp/providers/{sp}/cdf/declarations.cdf')
                data_raw = z.read(f'trp/providers/{sp}/cdf/data.cdf')
            except KeyError:
                continue

            decl_buf = decompress_cdf(decl_raw)
            data_buf = decompress_cdf(data_raw)

            for rec in iter_records(decl_buf):
                pid_entries = rec.get(2, [])
                name_entries = rec.get(1, [])
                if pid_entries and name_entries:
                    pid = pid_entries[0][1]
                    name_val = name_entries[0][1]
                    name_str = name_val.decode('utf-8', errors='replace') if isinstance(name_val, (bytes, bytearray)) else str(name_val)
                    all_param_names[pid] = name_str

            current_ts = None
            current_values = {}

            def flush():
                nonlocal current_ts, current_values
                if current_ts is None:
                    return
                lat = current_values.get(PID['LATITUDE'])
                lon = current_values.get(PID['LONGITUDE'])
                if lat and lon and lat != 0 and lon != 0:
                    ts_dt = datetime.fromtimestamp(current_ts, tz=timezone.utc)
                    rsrp = current_values.get(PID['RSRP'])
                    sinr = current_values.get(PID['SINR'])
                    dl = current_values.get(PID['PDSCH_TOTAL']) or current_values.get(PID['PDSCH_TP'])
                    ul = current_values.get(PID['PUSCH_TOTAL']) or current_values.get(PID['PUSCH_TP'])
                    samples.append({
                        'timestamp': ts_dt.strftime('%Y-%m-%d %H:%M:%S'),
                        'latitude': round(lat, 6),
                        'longitude': round(lon, 6),
                        'rsrp_dBm': round(rsrp, 1) if rsrp is not None else None,
                        'rsrq_dB': round(current_values.get(PID['RSRQ']), 1) if current_values.get(PID['RSRQ']) is not None else None,
                        'sinr_dB': round(sinr, 1) if sinr is not None else None,
                        'pci': current_values.get(PID['PCI']),
                        'band': current_values.get(PID['BAND']),
                        'dl_earfcn': current_values.get(PID['DL_EARFCN']),
                        'dl_throughput_kbps': round(dl, 1) if dl is not None else None,
                        'ul_throughput_kbps': round(ul, 1) if ul is not None else None,
                        'dl_throughput_mbps': round(dl / 1000, 2) if dl is not None else None,
                        'ul_throughput_mbps': round(ul / 1000, 2) if ul is not None else None,
                        'speed_kmh': round(current_values.get(PID['SPEED']), 1) if current_values.get(PID['SPEED']) is not None else None,
                        'altitude_m': round(current_values.get(PID['ALTITUDE']), 1) if current_values.get(PID['ALTITUDE']) is not None else None,
                        'coverage_class': rsrp_class(rsrp, cfg),
                        'sinr_class': sinr_class(sinr, cfg),
                    })

            for rec in iter_records(data_buf):
                ts = None
                ts_sub = rec.get(1, [])
                if ts_sub and ts_sub[0][0] == 'b':
                    inner = parse_pb(ts_sub[0][1])
                    if 1 in inner:
                        ts = inner[1][0][1]
                if ts and ts != current_ts:
                    flush()
                    current_ts = ts
                    current_values = {}
                for dp_type, dp_val in rec.get(3, []):
                    if dp_type != 'b':
                        continue
                    sub = parse_pb(dp_val)
                    pid_entries = sub.get(1, [])
                    if not pid_entries:
                        continue
                    pid = pid_entries[0][1]
                    val = extract_value(sub)
                    if val is not None:
                        current_values[pid] = val
            flush()

    samples.sort(key=lambda s: s['timestamp'])

    distance_km = 0
    for i in range(1, len(samples)):
        distance_km += haversine(
            samples[i-1]['latitude'], samples[i-1]['longitude'],
            samples[i]['latitude'], samples[i]['longitude'])

    return {
        'meta': meta,
        'param_names': all_param_names,
        'samples': samples,
        'total_samples': len(samples),
        'distance_km': round(distance_km, 2),
    }


# ---------- Excel output ----------

def save_to_excel(result, output_path, cfg=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    if cfg is None:
        cfg = DEFAULT_CONFIG
    wb = Workbook()
    meta = result['meta']
    samples = result['samples']

    # ---- Colors ----
    HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
    HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1F4E79')
    LABEL_FONT = Font(name='Arial', bold=True, size=10)
    VALUE_FONT = Font(name='Arial', size=10)
    DATA_FONT = Font(name='Arial', size=9)
    THIN_BORDER = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'))

    COVERAGE_FILLS = {
        'Excellent': PatternFill('solid', fgColor='C6EFCE'),
        'Good':      PatternFill('solid', fgColor='D9EAD3'),
        'Fair':      PatternFill('solid', fgColor='FFF2CC'),
        'Poor':      PatternFill('solid', fgColor='FCE4D6'),
        'No Signal': PatternFill('solid', fgColor='F4CCCC'),
    }
    SINR_FILLS = {
        'Excellent': PatternFill('solid', fgColor='C6EFCE'),
        'Good':      PatternFill('solid', fgColor='D9EAD3'),
        'Fair':      PatternFill('solid', fgColor='FFF2CC'),
        'Poor':      PatternFill('solid', fgColor='F4CCCC'),
    }

    # ==== Sheet 1: Summary ====
    ws1 = wb.active
    ws1.title = 'Summary'
    ws1.sheet_properties.tabColor = '1F4E79'

    ws1['A1'] = 'TEMS Drive Test Report'
    ws1['A1'].font = TITLE_FONT
    ws1.merge_cells('A1:D1')

    info = [
        ('Operator', meta['operator'] or 'Unknown'),
        ('MCC / MNC', f"{meta['mcc']} / {meta['mnc']}"),
        ('Device', meta['device'] or 'Unknown'),
        ('OS', meta['os'] or 'Unknown'),
        ('App Version', meta['app_version'] or 'Unknown'),
        ('IMEI', meta['imei'] or 'Unknown'),
        ('Tags', meta['tags'] or '-'),
        ('', ''),
        ('Total Samples', f"{result['total_samples']:,}"),
        ('Distance', f"{result['distance_km']} km"),
        ('Start Time', samples[0]['timestamp'] if samples else '-'),
        ('End Time', samples[-1]['timestamp'] if samples else '-'),
        ('', ''),
        ('--- KPI Thresholds (from DT Config) ---', ''),
        ('RSRP Threshold', f"{cfg['rsrp_threshold']} dBm"),
        ('RSRQ Threshold', f"{cfg['rsrq_threshold']} dB"),
        ('SINR Threshold', f"{cfg['sinr_threshold']} dB"),
        ('DL Threshold', f"{cfg['dl_threshold']} kbps"),
        ('UL Threshold', f"{cfg['ul_threshold']} kbps"),
        ('Coverage Target', f"{cfg['coverage_target']}%"),
        ('RSRP Classes', f"Excellent>={cfg['rsrp_excellent']}  Good>={cfg['rsrp_good']}  Fair>={cfg['rsrp_fair']}  Poor>={cfg['rsrp_poor']}"),
        ('SINR Classes', f"Excellent>={cfg['sinr_excellent']}  Good>={cfg['sinr_good']}  Fair>={cfg['sinr_fair']}"),
    ]

    for i, (label, value) in enumerate(info, start=3):
        ws1.cell(row=i, column=1, value=label).font = LABEL_FONT
        ws1.cell(row=i, column=2, value=value).font = VALUE_FONT

    # KPI stats
    row = len(info) + 5
    ws1.cell(row=row, column=1, value='KPI Summary').font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    row += 1
    kpi_headers = ['KPI', 'Average', 'Min', 'Max', 'Samples', 'Unit']
    for c, h in enumerate(kpi_headers, 1):
        cell = ws1.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    def kpi_row(name, key, unit, divisor=1):
        vals = [s[key] for s in samples if s[key] is not None]
        if not vals:
            return [name, '-', '-', '-', 0, unit]
        return [name, round(sum(vals)/len(vals)/divisor, 1), round(min(vals)/divisor, 1),
                round(max(vals)/divisor, 1), len(vals), unit]

    kpis = [
        kpi_row('RSRP', 'rsrp_dBm', 'dBm'),
        kpi_row('RSRQ', 'rsrq_dB', 'dB'),
        kpi_row('SINR', 'sinr_dB', 'dB'),
        kpi_row('DL Throughput', 'dl_throughput_kbps', 'Mbps', 1000),
        kpi_row('UL Throughput', 'ul_throughput_kbps', 'Mbps', 1000),
        kpi_row('Speed', 'speed_kmh', 'km/h'),
    ]
    for k in kpis:
        row += 1
        for c, v in enumerate(k, 1):
            cell = ws1.cell(row=row, column=c, value=v)
            cell.font = VALUE_FONT
            cell.border = THIN_BORDER
            if c > 1:
                cell.alignment = Alignment(horizontal='center')

    # Coverage distribution
    row += 2
    ws1.cell(row=row, column=1, value='Coverage Distribution (RSRP)').font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    row += 1
    for c, h in enumerate(['Class', 'Count', 'Percentage'], 1):
        cell = ws1.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    total = len(samples) or 1
    for cls in ['Excellent', 'Good', 'Fair', 'Poor', 'No Signal']:
        row += 1
        count = sum(1 for s in samples if s['coverage_class'] == cls)
        pct = round(count / total * 100, 1)
        ws1.cell(row=row, column=1, value=cls).font = VALUE_FONT
        ws1.cell(row=row, column=1).fill = COVERAGE_FILLS.get(cls, PatternFill())
        ws1.cell(row=row, column=2, value=count).font = VALUE_FONT
        ws1.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        c3 = ws1.cell(row=row, column=3, value=pct / 100)
        c3.font = VALUE_FONT
        c3.number_format = '0.0%'
        c3.alignment = Alignment(horizontal='center')
        for c in range(1, 4):
            ws1.cell(row=row, column=c).border = THIN_BORDER

    # SINR distribution
    row += 2
    ws1.cell(row=row, column=1, value='Signal Quality Distribution (SINR)').font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    row += 1
    for c, h in enumerate(['Class', 'Count', 'Percentage'], 1):
        cell = ws1.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    for cls in ['Excellent', 'Good', 'Fair', 'Poor']:
        row += 1
        count = sum(1 for s in samples if s['sinr_class'] == cls)
        pct = round(count / total * 100, 1)
        ws1.cell(row=row, column=1, value=cls).font = VALUE_FONT
        ws1.cell(row=row, column=1).fill = SINR_FILLS.get(cls, PatternFill())
        ws1.cell(row=row, column=2, value=count).font = VALUE_FONT
        ws1.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        c3 = ws1.cell(row=row, column=3, value=pct / 100)
        c3.font = VALUE_FONT
        c3.number_format = '0.0%'
        c3.alignment = Alignment(horizontal='center')
        for c in range(1, 4):
            ws1.cell(row=row, column=c).border = THIN_BORDER

    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 10

    # ==== Sheet 2: All Samples ====
    ws2 = wb.create_sheet('All Samples')
    ws2.sheet_properties.tabColor = '2E75B6'

    headers = [
        ('#', 5),
        ('Timestamp', 20),
        ('Latitude', 12),
        ('Longitude', 13),
        ('RSRP (dBm)', 12),
        ('RSRQ (dB)', 11),
        ('SINR (dB)', 11),
        ('PCI', 7),
        ('Band', 7),
        ('DL EARFCN', 11),
        ('DL (kbps)', 12),
        ('UL (kbps)', 12),
        ('DL (Mbps)', 11),
        ('UL (Mbps)', 11),
        ('Speed (km/h)', 13),
        ('Altitude (m)', 13),
        ('Coverage Class', 15),
        ('SINR Class', 12),
    ]

    for c, (h, w) in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws2.column_dimensions[get_column_letter(c)].width = w

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws2.freeze_panes = 'A2'

    fields = ['timestamp', 'latitude', 'longitude', 'rsrp_dBm', 'rsrq_dB', 'sinr_dB',
              'pci', 'band', 'dl_earfcn', 'dl_throughput_kbps', 'ul_throughput_kbps',
              'dl_throughput_mbps', 'ul_throughput_mbps', 'speed_kmh', 'altitude_m',
              'coverage_class', 'sinr_class']

    for r, sample in enumerate(samples, 2):
        ws2.cell(row=r, column=1, value=r - 1).font = DATA_FONT
        ws2.cell(row=r, column=1).alignment = Alignment(horizontal='center')
        for c, key in enumerate(fields, 2):
            val = sample[key]
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            if key == 'coverage_class' and val:
                cell.fill = COVERAGE_FILLS.get(val, PatternFill())
                cell.alignment = Alignment(horizontal='center')
            elif key == 'sinr_class' and val:
                cell.fill = SINR_FILLS.get(val, PatternFill())
                cell.alignment = Alignment(horizontal='center')
            elif isinstance(val, float):
                cell.number_format = '0.0' if abs(val) < 1000 else '#,##0.0'
                cell.alignment = Alignment(horizontal='right')
            elif isinstance(val, int):
                cell.alignment = Alignment(horizontal='center')

    # ==== Sheet 3: Parameters ====
    ws3 = wb.create_sheet('Parameters')
    ws3.sheet_properties.tabColor = '7030A0'

    for c, (h, w) in enumerate([('PID', 10), ('Parameter Name', 55), ('Used In Parser', 15)], 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        ws3.column_dimensions[get_column_letter(c)].width = w

    ws3.auto_filter.ref = 'A1:C1'
    ws3.freeze_panes = 'A2'

    PID_REV = {v: k for k, v in PID.items()}
    for r, (pid, name) in enumerate(sorted(result['param_names'].items()), 2):
        ws3.cell(row=r, column=1, value=pid).font = DATA_FONT
        ws3.cell(row=r, column=1).alignment = Alignment(horizontal='center')
        ws3.cell(row=r, column=2, value=name).font = DATA_FONT
        used = PID_REV.get(pid, '')
        cell = ws3.cell(row=r, column=3, value=used)
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal='center')
        if used:
            cell.fill = PatternFill('solid', fgColor='C6EFCE')

    wb.save(output_path)
    print(f"\n  Excel saved: {output_path}")
    print(f"  - Summary sheet: metadata + KPI stats + coverage distribution")
    print(f"  - All Samples sheet: {len(samples):,} rows with filters and color coding")
    print(f"  - Parameters sheet: {len(result['param_names']):,} declared parameters\n")


# ---------- Run ----------

if __name__ == '__main__':
    filepath = Path(TRP_FILE)
    if not filepath.exists():
        print(f"ERROR: File not found: {TRP_FILE}")
        print(f"Please update TRP_FILE at the top of the script.")
        exit(1)

    cfg = fetch_config()
    print(f"\n  Thresholds: RSRP>={cfg['rsrp_threshold']}dBm  SINR>={cfg['sinr_threshold']}dB  DL>={cfg['dl_threshold']}kbps")
    print(f"  Coverage classes: Excellent>={cfg['rsrp_excellent']}  Good>={cfg['rsrp_good']}  Fair>={cfg['rsrp_fair']}  Poor>={cfg['rsrp_poor']}")

    print(f"\nParsing: {filepath.name}")
    result = parse_trp(str(filepath), cfg)

    meta = result['meta']
    print(f"\n  Operator:  {meta['operator'] or 'Unknown'}")
    print(f"  Device:    {meta['device'] or 'Unknown'}")
    print(f"  Samples:   {result['total_samples']:,}")
    print(f"  Distance:  {result['distance_km']} km")

    save_to_excel(result, OUTPUT_EXCEL, cfg)
    print("Done!")
